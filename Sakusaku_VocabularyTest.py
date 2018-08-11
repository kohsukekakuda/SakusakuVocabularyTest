# coding: utf-8

import sys
import random
import csv
import argparse
import os
import openpyxl

def choice(read):
    Input = (random.sample(read, int(args.cnt)))
    return Input
    
def write_txt(read, inputcsv):
    cd = os.getcwd()
    L = choice(read)
    LL = []
    for i in xrange(int(args.cnt)):
        LL.append(L[i].replace("\n", ""))
    Z = [LL[i:i+1] for i in xrange(0, len(LL), 1)]
    Z = [Z[i:i+2] for i in xrange(0, len(Z), 2)]
    Z = [z[::2] + z[1::2] for z in Z]

    wb = openpyxl.Workbook()
    ws = wb.active

    ws["A1"] = ("%s%s%s%s\n\n" % ("EnglishVocabularyTest Target "
                             , int(inputcsv[32:38]) - int(99)
                             , " ~ "
                             , int(inputcsv[32:38])))
    AAA = []
    BBB = []
    for xs in Z:
        n = (len(xs) + 1) // 2
        list1 = ' '.join(sum(xs[:n], []))
        list2 = ' '.join(sum(xs[n:], []))
        AAA.append(list1)
        BBB.append(list2)
    AAA.extend(BBB)
    AAA = sorted(AAA)

    if int(args.cnt) % 2:
        del AAA[0]
    else:
        pass

    if int(args.cnt) <= 50:
        for i in xrange(int(args.cnt)):
            if (i+2)*2 < 54:
                ws.cell(row=(i+2)*2, column=1).value = AAA[i]
            elif 54 <= (i+2)*2 < 104:
                ws.cell(row=(i-23)*2, column=5).value = AAA[i]
    else:
        for i in xrange(int(args.cnt)):
            if (i+2)*2 < 54:
                ws.cell(row=(i+2)*2, column=1).value = AAA[i]
            elif 54 <= (i+2)*2 < 104:
                ws.cell(row=(i-23)*2, column=3).value = AAA[i]
            elif 104 <= (i+2)*2 < 154:
                ws.cell(row=(i-48)*2, column=5).value = AAA[i]
            elif 154 <= (i+2)*2:
                ws.cell(row=(i-73)*2, column=7).value = AAA[i]

    wb.save(cd + "/_VocabularyTest/" + "VocabularyTest_nm"
            + ("%s" % (int(args.mode[32:38]) - int(99))) 
            + "-"
            + ("%s" % (int(args.mode[32:38])))
            + ".xlsx")

def read_csv(fp):
    data = []
    for line in fp:
        data.append(line)
    return data

def main():
    cd = os.getcwd()
    if os.path.exists(cd + "/_VocabularyTest/") == False:
        os.mkdir(cd + "/_VocabularyTest")
    elif os.path.exists(cd + "/_VocabularyTest") == True:
        pass
    with open(args.mode) as fp:
        read = read_csv(fp)
        inputcsv = args.mode
    write_txt(read, inputcsv)
    if len(sys.argv) < 5:
        print "Usage: python Sakusaku_VocabularyTest.py -s VocabularyBook/VocabularyBook_nm000100.csv -c 10"
        sys.exit()
    else:
        print ("Out: " + "VocabularyTest_nm"
               + ("%s" % (int(args.mode[32:38]) - int(99))) 
               + "-"
               + ("%s" % (int(args.mode[32:38])))
               + ".xlsx")
        
if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='')
    parser.add_argument('-s', metavar='Select csv-file', dest='mode', type=str, default='VocabularyBook/VocabularyBook_nm000100.csv',
                        help='')
    parser.add_argument('-c', metavar='Select number', dest='cnt', type=str, default='25',
                        help='')
    args = parser.parse_args()
    main()
    
